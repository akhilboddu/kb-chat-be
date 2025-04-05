import io
import csv
from typing import Optional
from fastapi import UploadFile
import pypdf  # Changed from PyPDF2 to pypdf
import docx # python-docx
import openpyxl

SUPPORTED_EXTENSIONS = {".txt", ".md", ".csv", ".pdf", ".docx", ".xlsx"}

def get_file_extension(filename: str) -> Optional[str]:
    """Extracts the file extension in lowercase."""
    parts = filename.rsplit('.', 1)
    if len(parts) > 1:
        return f".{parts[-1].lower()}"
    return None

async def _parse_txt_md(file: UploadFile) -> str:
    """Parses text from TXT or MD files."""
    content_bytes = await file.read()
    return content_bytes.decode('utf-8') # Assume UTF-8, could add error handling

async def _parse_csv(file: UploadFile) -> str:
    """Parses text from CSV files, joining rows and cells."""
    content_bytes = await file.read()
    text_content = []
    try:
        # Use StringIO to treat bytes as a text file
        file_like_object = io.StringIO(content_bytes.decode('utf-8')) # Assume UTF-8
        reader = csv.reader(file_like_object)
        for row in reader:
            text_content.append(", ".join(cell.strip() for cell in row if cell))
        return "\n".join(text_content)
    except Exception as e:
        print(f"Error parsing CSV: {e}")
        # Fallback: return raw text if CSV parsing fails
        return content_bytes.decode('utf-8', errors='ignore')

async def _parse_pdf(file: UploadFile) -> str:
    """Parses text from PDF files."""
    content_bytes = await file.read()
    text_content = []
    try:
        pdf_reader = pypdf.PdfReader(io.BytesIO(content_bytes))  # Changed from PyPDF2.PdfReader to pypdf.PdfReader
        for page in pdf_reader.pages:
            page_text = page.extract_text()
            if page_text:
                text_content.append(page_text)
        return "\n".join(text_content)
    except Exception as e:
        print(f"Error parsing PDF: {e}")
        return "" # Return empty string on error

async def _parse_docx(file: UploadFile) -> str:
    """Parses text from DOCX files."""
    content_bytes = await file.read()
    text_content = []
    try:
        document = docx.Document(io.BytesIO(content_bytes))
        for para in document.paragraphs:
            if para.text:
                text_content.append(para.text.strip())
        # Consider adding table content extraction if needed
        return "\n".join(text_content)
    except Exception as e:
        print(f"Error parsing DOCX: {e}")
        return "" # Return empty string on error

async def _parse_xlsx(file: UploadFile) -> str:
    """Parses text from XLSX files, joining cells."""
    content_bytes = await file.read()
    text_content = []
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(content_bytes), data_only=True) # data_only=True to get values, not formulas
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                row_text = []
                for cell in row:
                    if cell.value is not None:
                        row_text.append(str(cell.value).strip())
                if row_text:
                    text_content.append(", ".join(row_text))
        return "\n".join(text_content)
    except Exception as e:
        print(f"Error parsing XLSX: {e}")
        return "" # Return empty string on error

async def parse_file(file: UploadFile) -> Optional[str]:
    """
    Parses the uploaded file based on its extension and returns the extracted text.

    Args:
        file: The uploaded file object from FastAPI.

    Returns:
        The extracted text content as a string, or None if the file type
        is unsupported or parsing fails significantly.
    """
    if not file.filename:
        print("Error: Uploaded file has no filename.")
        return None

    extension = get_file_extension(file.filename)
    if not extension or extension not in SUPPORTED_EXTENSIONS:
        print(f"Error: Unsupported file extension '{extension}' for file '{file.filename}'. Supported: {SUPPORTED_EXTENSIONS}")
        return None

    print(f"Parsing file: {file.filename} (Extension: {extension})")

    try:
        if extension in (".txt", ".md"):
            return await _parse_txt_md(file)
        elif extension == ".csv":
            return await _parse_csv(file)
        elif extension == ".pdf":
            return await _parse_pdf(file)
        elif extension == ".docx":
            return await _parse_docx(file)
        elif extension == ".xlsx":
            return await _parse_xlsx(file)
        else:
            # Should not happen due to earlier check, but as a safeguard
            print(f"Warning: Reached unexpected parsing path for extension {extension}")
            return None
    except Exception as e:
        # Catch any unexpected errors during the specific parsing functions
        print(f"Critical error during parsing of {file.filename}: {e}")
        import traceback
        traceback.print_exc()
        return None
    finally:
         # Ensure the file pointer is reset if necessary, though read() should consume it
         # await file.seek(0) # Usually not needed with await file.read()
         pass

# Example usage placeholder (won't run directly here)
# async def main():
#     # This requires mocking UploadFile, typically tested via endpoint
#     pass
# if __name__ == "__main__":
#    # asyncio.run(main())
#    pass 